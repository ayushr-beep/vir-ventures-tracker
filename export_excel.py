"""
Vir Ventures — 6-sheet premium Excel workbook with Vendor Analysis.
"""
import io, numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from datetime import datetime

NV="0F1629"; NV2="1E2A45"; OR="E8611A"; OR2="F4874B"
WH="FFFFFF"; LT="F8F9FB"; LG="E5E7EB"; GY="6B7280"
GR="16A34A"; RD="DC2626"; AM="D97706"; BL="2563EB"
DK="111827"; LGR="DCFCE7"; LRD="FEE2E2"; LAM="FEF9C3"; LOR="FFF0E8"
VENDOR_NAMES={"CRDI":"CRDI — Gaming & Entertainment","PNCL":"PNCL — The Pencil Grip","PROH":"PROH — Prohitter Sports"}
VENDOR_COLORS_XL={"CRDI":OR,"PNCL":BL,"PROH":AM}

YES_VALS={"yes","yes, need discount"}
def is_yes(v): return str(v).strip().lower() in YES_VALS

def nf(h): return PatternFill("solid",fgColor=h)
def fw(s=10,b=True): return Font(name="Calibri",bold=b,size=s,color=WH)
def fd(s=10,b=False,c=DK): return Font(name="Calibri",bold=b,size=s,color=c)
def fo(s=10): return Font(name="Calibri",bold=True,size=s,color=OR)
def fg(s=10): return Font(name="Calibri",bold=True,size=s,color=GR)
def fr(s=10): return Font(name="Calibri",bold=True,size=s,color=RD)
def fa(s=10): return Font(name="Calibri",bold=True,size=s,color=AM)
def fl(s=10): return Font(name="Calibri",bold=True,size=s,color=BL)

def cen(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
def lft(): return Alignment(horizontal="left",vertical="center",wrap_text=True)

def th(c=LG):
    s=Side(style="thin",color=c)
    return Border(left=s,right=s,top=s,bottom=s)

def rate_style(r):
    if r>=80: return nf(LGR), Font(name="Calibri",bold=True,size=10,color=GR)
    if r>=50: return nf(LAM), Font(name="Calibri",bold=True,size=10,color=AM)
    return nf(LRD), Font(name="Calibri",bold=True,size=10,color=RD)

def sg(ws): ws.sheet_view.showGridLines=False

def banner(ws,r,text,cols,h=30):
    ws.merge_cells(f"A{r}:{get_column_letter(cols)}{r}")
    c=ws.cell(row=r,column=1,value=text)
    c.font=Font(name="Calibri",bold=True,size=13,color=WH)
    c.fill=nf(NV); c.alignment=cen()
    ws.row_dimensions[r].height=h

def sub_hdr(ws,r,text,cols,h=18):
    ws.merge_cells(f"A{r}:{get_column_letter(cols)}{r}")
    c=ws.cell(row=r,column=1,value=text)
    c.font=Font(name="Calibri",bold=True,size=9,color=OR)
    c.fill=nf(NV2); c.alignment=lft()
    ws.row_dimensions[r].height=h

def col_hdr(ws,r,cols_list,start=1):
    for i,(lbl,w) in enumerate(cols_list):
        c=ws.cell(row=r,column=start+i,value=lbl)
        c.font=fw(9); c.fill=nf(NV2); c.alignment=cen(); c.border=th()
        ws.column_dimensions[get_column_letter(start+i)].width=w
    ws.row_dimensions[r].height=20

def data_cell(ws,r,col,val,bg=WH,font=None,align=None):
    c=ws.cell(row=r,column=col,value=val)
    c.fill=nf(bg); c.border=th()
    c.font=font or fd(10)
    c.alignment=align or cen()
    return c

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 — EXECUTIVE DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════
def sh_exec(wb,d):
    ws=wb.create_sheet("📊 Executive Dashboard"); sg(ws)
    df=d['df']; yl=df[df['is_yes']] if 'is_yes' in df.columns else df[df['Recommended'].apply(is_yes)]

    banner(ws,1,f"VIR VENTURES — BOARDROOM ANALYST DASHBOARD  ·  {d['week_label'].upper()}",12,34)
    ws.merge_cells("A2:L2")
    c=ws.cell(row=2,column=1,value=f"Generated {datetime.now().strftime('%d %b %Y %H:%M')}  ·  Confidential  ·  {d['report_date']}")
    c.font=Font(name="Calibri",size=8,color=OR2); c.fill=nf(NV2); c.alignment=cen()
    ws.row_dimensions[2].height=14

    # KPI row
    kpis=[("CATALOGUE SIZE",str(d['total']),"Total SKUs Reviewed",OR,DK),
          ("APPROVED",str(d['approved']),f"{d['rate']}% approval rate",GR,GR),
          ("REJECTED",str(d['rejected']),"Negative/Low Margin",RD,RD),
          ("APPROVAL RATE",f"{d['rate']}%",f"of {d['total']} SKUs",BL,BL)]
    r=4
    for i,(lbl,val,sub,acc,vc) in enumerate(kpis):
        c0=1+i*3
        ws.merge_cells(start_row=r,end_row=r,start_column=c0,end_column=c0+1)
        ws.merge_cells(start_row=r+1,end_row=r+2,start_column=c0,end_column=c0+1)
        ws.merge_cells(start_row=r+3,end_row=r+3,start_column=c0,end_column=c0+1)
        for ri2,v2,fs,fc in [(r,lbl,8,GY),(r+1,val,22,vc),(r+3,sub,8,GY)]:
            cc=ws.cell(row=ri2,column=c0,value=v2)
            cc.font=Font(name="Calibri",bold=(ri2==r+1),size=fs,color=fc)
            cc.fill=nf(WH); cc.alignment=lft() if ri2!=r+1 else cen()
        ws.row_dimensions[r].height=13; ws.row_dimensions[r+1].height=24
        ws.row_dimensions[r+2].height=8; ws.row_dimensions[r+3].height=13
    ws.row_dimensions[r+4].height=10

    # Highlights
    r2=r+5
    top_a=df.groupby('Analyst')['is_yes'].mean().idxmax() if not yl.empty else "—"
    top_ar=round(df.groupby('Analyst')['is_yes'].mean().max()*100,1) if not yl.empty else 0
    top_brand=yl.groupby('Brand')['ASIN'].count().idxmax() if not yl.empty else "—"
    top_brand_n=int(yl.groupby('Brand')['ASIN'].count().max()) if not yl.empty else 0
    top_asin_row=yl.sort_values('Last month sold',ascending=False).iloc[0] if not yl.empty else None
    avg_m=round(yl['margin_pct'].mean(),1) if not yl.empty else 0
    hi=[("TOP ANALYST",top_a,f"{top_ar}% Approval Rate"),
        ("TOP BRAND",top_brand,f"{top_brand_n} Approved SKUs"),
        ("TOP ASIN",top_asin_row['ASIN'] if top_asin_row is not None else "—",
         f"{int(top_asin_row['Last month sold'])} units/mo" if top_asin_row is not None else "—"),
        ("AVG MARGIN",f"{avg_m}%","On approved SKUs")]
    for i,(lbl,val,sub) in enumerate(hi):
        c0=1+i*3
        ws.merge_cells(start_row=r2,end_row=r2,start_column=c0,end_column=c0+1)
        ws.merge_cells(start_row=r2+1,end_row=r2+2,start_column=c0,end_column=c0+1)
        ws.merge_cells(start_row=r2+3,end_row=r2+3,start_column=c0,end_column=c0+1)
        for ri2,v2,fs,fc in [(r2,lbl,8,OR),(r2+1,val,15,DK),(r2+3,sub,8,GY)]:
            cc=ws.cell(row=ri2,column=c0,value=v2)
            cc.font=Font(name="Calibri",bold=(ri2==r2+1),size=fs,color=fc)
            cc.fill=nf(LOR if ri2==r2 else WH); cc.alignment=lft() if ri2!=r2+1 else cen()
        ws.row_dimensions[r2].height=13; ws.row_dimensions[r2+1].height=20
        ws.row_dimensions[r2+2].height=6; ws.row_dimensions[r2+3].height=13
    ws.row_dimensions[r2+4].height=10

    # Analyst summary
    r3=r2+5
    sub_hdr(ws,r3,"  ANALYST PERFORMANCE SUMMARY",12)
    r3+=1
    col_hdr(ws,r3,[("Analyst",14),("Total",11),("Approved",11),("Rejected",11),
                    ("Rate",12),("Avg Margin %",13),("Units/Mo",12),("Avg Rank",11),
                    ("Demand Sc.",11),("Comp Sc.",11),("Margin Sc.",11),("Avg BB $",11)])
    r3+=1
    astats=df.groupby('Analyst').agg(
        Total=('ASIN','count'),Approved=('is_yes','sum'),
        Rejected=('is_yes',lambda x:(~x).sum()),
        Rate=('is_yes',lambda x:round(x.mean()*100,1)),
        AvgM=('margin_pct','mean'),Units=('Last month sold','sum'),
        Rank=('Rank','mean'),Dem=('Total Demand Score','mean'),
        Comp=('Total Competition Score','mean'),MarSc=('Total Margin Score','mean'),
        BB=('BB Price','mean')).reset_index().round(1)
    for ri,(_,row) in enumerate(astats.iterrows()):
        bg=WH if ri%2==0 else LT
        rf,rfont=rate_style(row['Rate'])
        vals=[row['Analyst'],int(row['Total']),int(row['Approved']),int(row['Rejected']),
              row['Rate'],row['AvgM'],int(row['Units']),round(row['Rank']),
              row['Dem'],row['Comp'],row['MarSc'],round(row['BB'],2)]
        for ci,v in enumerate(vals,1):
            c=ws.cell(row=r3,column=ci,value=v)
            c.fill=rf if ci==5 else nf(bg)
            c.font=rfont if ci==5 else fd(10,b=(ri==0 and ci==1))
            c.alignment=cen(); c.border=th()
        ws.row_dimensions[r3].height=19; r3+=1

    ws.freeze_panes="A9"
    for col_l,w in [("A",14),("B",11),("C",11),("D",11),("E",12),("F",13),
                     ("G",12),("H",11),("I",11),("J",11),("K",11),("L",11)]:
        ws.column_dimensions[col_l].width=w

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 — VENDOR ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════
def sh_vendor(wb,d):
    ws=wb.create_sheet("🏪 Vendor Analysis"); sg(ws)
    df=d['df']
    if 'Vendor Prefix' not in df.columns:
        ws.cell(row=1,column=1,value="No Vendor Prefix column found.")
        return

    banner(ws,1,"VENDOR ANALYSIS — PERFORMANCE BY VENDOR GROUP",11,34)
    ws.merge_cells("A2:K2")
    c=ws.cell(row=2,column=1,value=f"Vendor breakdown for {d['week_label']}  ·  {d['report_date']}")
    c.font=Font(name="Calibri",size=8,color=OR2); c.fill=nf(NV2); c.alignment=cen()
    ws.row_dimensions[2].height=14; ws.row_dimensions[3].height=10

    vp=df.groupby('Vendor Prefix').agg(
        Total=('ASIN','count'),Approved=('is_yes','sum'),
        Rejected=('is_yes',lambda x:(~x).sum()),
        Brands=('Brand','nunique'),
        Units=('Last month sold','sum'),
        AvgM=('margin_pct','mean'),AvgRank=('Rank','mean'),
        AvgNet=('Net price','mean'),AvgBB=('BB Price','mean'),
        AvgDem=('Total Demand Score','mean'),
        AvgComp=('Total Competition Score','mean'),
        AvgMarSc=('Total Margin Score','mean'),
    ).reset_index().round(1)
    vp['Rate']=(vp['Approved']/vp['Total']*100).round(1)
    vp['Name']=vp['Vendor Prefix'].map(VENDOR_NAMES).fillna(vp['Vendor Prefix'])

    # Vendor summary cards (row 4)
    r=4
    sub_hdr(ws,r,"  VENDOR SCORECARD SUMMARY",11); r+=1
    col_hdr(ws,r,[("Vendor",8),("Vendor Name",28),("Total SKUs",11),("Approved",11),
                   ("Rejected",11),("Approval Rate",13),("Total Units/Mo",14),("Avg Margin %",13),
                   ("Avg BSR Rank",12),("Avg Net Price",12),("Avg BB Price",12)])
    r+=1
    for ri,(_,row) in enumerate(vp.iterrows()):
        vc_h=VENDOR_COLORS_XL.get(row['Vendor Prefix'],OR)
        bg=LOR if ri==0 else (nf(WH).fgColor.rgb if ri%2==0 else LT)
        rf,rfont=rate_style(row['Rate'])
        vals=[row['Vendor Prefix'],row['Name'],int(row['Total']),int(row['Approved']),
              int(row['Rejected']),row['Rate'],int(row['Units']),row['AvgM'],
              round(row['AvgRank']),round(row['AvgNet'],2),round(row['AvgBB'],2)]
        for ci,v in enumerate(vals,1):
            c=ws.cell(row=r,column=ci,value=v)
            if ci==1: c.fill=nf(vc_h); c.font=fw(10,True)
            elif ci==6: c.fill=rf; c.font=rfont
            elif ci==8: c.fill=nf(LOR); c.font=fo(10)
            else: c.fill=nf(LOR if ri==0 else (WH if ri%2==0 else LT)); c.font=fd(10,b=(ri==0 and ci==2))
            c.alignment=cen(); c.border=th()
        ws.row_dimensions[r].height=20; r+=1

    r+=1
    # Bar chart
    chart_s=6; chart_e=5+len(vp)
    cht=BarChart(); cht.type="col"; cht.title="Approved vs Rejected by Vendor"
    cht.style=10; cht.width=16; cht.height=10
    d1=Reference(ws,min_col=4,min_row=chart_s,max_row=chart_e)
    d2=Reference(ws,min_col=5,min_row=chart_s,max_row=chart_e)
    ct=Reference(ws,min_col=1,min_row=chart_s,max_row=chart_e)
    cht.add_data(d1,titles_from_data=False)
    cht.add_data(d2,titles_from_data=False)
    cht.set_categories(ct)
    cht.series[0].graphicalProperties.solidFill=GR
    cht.series[1].graphicalProperties.solidFill=RD
    ws.add_chart(cht,"A"+str(r))
    r+=17

    # Per-vendor deep dive
    for _,vrow in vp.iterrows():
        pfx=vrow['Vendor Prefix']
        vc_h=VENDOR_COLORS_XL.get(pfx,OR)
        vdf=df[df['Vendor Prefix']==pfx].sort_values('Last month sold',ascending=False)
        vname=VENDOR_NAMES.get(pfx,pfx)

        # Vendor header
        ws.merge_cells(f"A{r}:K{r}")
        c=ws.cell(row=r,column=1,value=f"  {pfx}  —  {vname}")
        c.font=Font(name="Calibri",bold=True,size=11,color=WH)
        c.fill=nf(vc_h); c.alignment=lft()
        ws.row_dimensions[r].height=22; r+=1

        # Metrics row
        vmet=[("Total SKUs",int(vrow['Total'])),("Approved",int(vrow['Approved'])),
              ("Rejected",int(vrow['Rejected'])),("Approval Rate",f"{vrow['Rate']}%"),
              ("Avg Margin",f"{vrow['AvgM']:.1f}%"),("Units/Month",int(vrow['Units'])),
              ("Brands",int(vrow['Brands'])),("Avg BSR Rank",round(vrow['AvgRank']))]
        for ci2,(lbl,val) in enumerate(vmet,1):
            cL=ws.cell(row=r,column=ci2,value=lbl)
            cL.font=fd(8,c=GY); cL.fill=nf(LT); cL.alignment=cen(); cL.border=th()
            cV=ws.cell(row=r+1,column=ci2,value=val)
            cV.font=fo(11) if lbl in ("Avg Margin","Approval Rate") else fd(11,b=True)
            cV.fill=nf(WH); cV.alignment=cen(); cV.border=th()
        ws.row_dimensions[r].height=16; ws.row_dimensions[r+1].height=20; r+=2

        # ASIN table
        col_hdr(ws,r,[("ASIN",14),("Brand",20),("Analyst",11),("Recommended",16),
                       ("Units/Mo",11),("Net $",10),("BB $",10),("Margin%",10),
                       ("BSR Rank",11),("Score",9),("BB Holder",16),("Remarks",28)])
        r+=1
        for ri2,(_,row2) in enumerate(vdf.iterrows()):
            bg=LGR if row2['is_yes'] else LRD
            vals=[row2.get('ASIN',''),row2.get('Brand',''),row2.get('Analyst',''),
                  row2.get('Recommended',''),int(row2.get('Last month sold',0)),
                  round(row2.get('Net price',0),2),round(row2.get('BB Price',0),2),
                  round(row2.get('margin_pct',0),1),int(row2.get('Rank',0)),
                  row2.get('Total Product Score',''),
                  str(row2.get('BuyBoxSellerName',''))[:16],str(row2.get('Remarks ',''))[:35]]
            for ci2,v in enumerate(vals,1):
                c=ws.cell(row=r,column=ci2,value=v)
                c.fill=nf(LGR if row2['is_yes'] else LRD)
                c.font=Font(name="Calibri",size=9,
                            color=GR if (row2['is_yes'] and ci2==4) else (RD if (not row2['is_yes'] and ci2==4) else DK))
                c.alignment=cen(); c.border=th()
            ws.row_dimensions[r].height=17; r+=1
        r+=2

    ws.freeze_panes="A6"

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3 — ASIN INTELLIGENCE
# ═══════════════════════════════════════════════════════════════════════════
def sh_asins(wb,d):
    ws=wb.create_sheet("📦 ASIN Intelligence"); sg(ws)
    df=d['df']
    yl=df[df['is_yes']].sort_values('Last month sold',ascending=False)

    banner(ws,1,"ASIN INTELLIGENCE — APPROVED SKUs RANKED BY DEMAND",15,32)
    ws.row_dimensions[2].height=8

    r=3
    sub_hdr(ws,r,f"  {len(yl)} APPROVED ASINs  ·  Ranked by monthly demand velocity",15); r+=1
    col_hdr(ws,r,[("#",4),("ASIN",14),("Vendor",8),("Brand",20),("Analyst",11),
                   ("Recommended",16),("Units/Mo",11),("Net $",10),("BB $",10),
                   ("Margin $",10),("Margin %",10),("BSR Rank",11),("Score",9),
                   ("FBA Sellers",11),("BB Holder",16),("Remarks",28)])
    r+=1
    for ri,(_,row) in enumerate(yl.iterrows()):
        u=int(row.get('Last month sold',0))
        bg="F0FDF4" if u>=200 else ("FFF7ED" if u>=100 else (LT if u>=50 else WH))
        vals=[ri+1,row.get('ASIN',''),row.get('Vendor Prefix',''),row.get('Brand',''),
              row.get('Analyst',''),row.get('Recommended',''),u,
              round(row.get('Net price',0),2),round(row.get('BB Price',0),2),
              round(row.get('margin_gap',0),2),round(row.get('margin_pct',0),1),
              int(row.get('Rank',0)),row.get('Total Product Score',''),
              int(row.get('Number of FBA vendors',0)),
              str(row.get('BuyBoxSellerName',''))[:16],str(row.get('Remarks ',''))[:35]]
        for ci,v in enumerate(vals,1):
            c=ws.cell(row=r,column=ci,value=v)
            c.fill=nf(bg); c.border=th(); c.alignment=cen()
            c.font=fd(9,b=(ri==0))
            if ci==7: c.font=Font(name="Calibri",bold=True,size=9,color=GR)
            if ci==11: c.font=fo(9)
        ws.row_dimensions[r].height=17; r+=1

    ws.freeze_panes="A5"
    ws.auto_filter.ref=f"A4:{get_column_letter(16)}{4+len(yl)}"

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4 — ANALYST SCORECARDS
# ═══════════════════════════════════════════════════════════════════════════
def sh_analysts(wb,d):
    ws=wb.create_sheet("👥 Analyst Scorecards"); sg(ws)
    df=d['df']
    banner(ws,1,"ANALYST PERFORMANCE SCORECARDS",10,32)
    ws.row_dimensions[2].height=8

    astats=df.groupby('Analyst').agg(
        Total=('ASIN','count'),Approved=('is_yes','sum'),
        Rejected=('is_yes',lambda x:(~x).sum()),
        Rate=('is_yes',lambda x:round(x.mean()*100,1)),
        AvgM=('margin_pct','mean'),Units=('Last month sold','sum'),
        Rank=('Rank','mean'),Net=('Net price','mean'),BB=('BB Price','mean'),
        Dem=('Total Demand Score','mean'),Comp=('Total Competition Score','mean'),
        MarSc=('Total Margin Score','mean')).reset_index().round(1)

    r=3
    for _,row in astats.iterrows():
        rate=row['Rate']
        acc=GR if rate>=80 else (AM if rate>=50 else RD)

        ws.merge_cells(f"A{r}:J{r}")
        c=ws.cell(row=r,column=1,value=f"  {row['Analyst'].upper()}  ·  {rate}% Approval Rate  ·  {int(row['Total'])} SKUs Audited")
        c.font=Font(name="Calibri",bold=True,size=12,color=WH)
        c.fill=nf(NV2); c.alignment=lft(); ws.row_dimensions[r].height=24; r+=1

        met=[("Total SKUs",int(row['Total'])),("Approved",int(row['Approved'])),
             ("Rejected",int(row['Rejected'])),("Approval Rate",f"{rate}%"),
             ("Avg Margin%",f"{row['AvgM']:.1f}%"),("Units/Month",int(row['Units'])),
             ("Avg BSR Rank",round(row['Rank'])),("Avg Net $",round(row['Net'],2)),
             ("Avg BB $",round(row['BB'],2)),("Demand Sc.",row['Dem'])]
        for ci2,(lbl,val) in enumerate(met,1):
            cL=ws.cell(row=r,column=ci2,value=lbl)
            cL.font=fd(8,c=GY); cL.fill=nf(LT); cL.alignment=cen(); cL.border=th()
            cV=ws.cell(row=r+1,column=ci2,value=val)
            fc=GR if lbl=="Approved" else (RD if lbl=="Rejected" else (OR if "Margin" in lbl or "Rate" in lbl else DK))
            cV.font=Font(name="Calibri",bold=True,size=11,color=fc)
            cV.fill=nf(WH); cV.alignment=cen(); cV.border=th()
        ws.row_dimensions[r].height=16; ws.row_dimensions[r+1].height=20; r+=2

        adf=df[df['Analyst']==row['Analyst']].sort_values('Last month sold',ascending=False)
        sub_hdr(ws,r,f"  {row['Analyst']} — ASIN DETAIL",10); r+=1
        col_hdr(ws,r,[("ASIN",14),("Vendor",8),("Brand",20),("Recommended",16),
                       ("Units/Mo",11),("Net $",10),("BB $",10),("Margin%",10),
                       ("BSR Rank",11),("Score",9),("Remarks",28)]); r+=1
        for ri2,(_,ar) in enumerate(adf.iterrows()):
            vals=[ar.get('ASIN',''),ar.get('Vendor Prefix',''),ar.get('Brand',''),ar.get('Recommended',''),
                  int(ar.get('Last month sold',0)),round(ar.get('Net price',0),2),
                  round(ar.get('BB Price',0),2),round(ar.get('margin_pct',0),1),
                  int(ar.get('Rank',0)),ar.get('Total Product Score',''),str(ar.get('Remarks ',''))[:35]]
            for ci2,v in enumerate(vals,1):
                c=ws.cell(row=r,column=ci2,value=v)
                c.fill=nf(LGR if ar['is_yes'] else LRD); c.border=th(); c.alignment=cen()
                c.font=Font(name="Calibri",size=9,
                            color=GR if (ar['is_yes'] and ci2==4) else (RD if (not ar['is_yes'] and ci2==4) else DK))
            ws.row_dimensions[r].height=17; r+=1
        r+=2
    ws.freeze_panes="A4"

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5 — BRAND INTELLIGENCE
# ═══════════════════════════════════════════════════════════════════════════
def sh_brands(wb,d):
    ws=wb.create_sheet("🏷️ Brand Intelligence"); sg(ws)
    df=d['df']
    banner(ws,1,"BRAND INTELLIGENCE — VENDOR PERFORMANCE & SOURCING PRIORITY",9,32)
    ws.row_dimensions[2].height=8

    bs=df.groupby('Brand').agg(
        Total=('ASIN','count'),Approved=('is_yes','sum'),
        Rejected=('is_yes',lambda x:(~x).sum()),
        Units=('Last month sold','sum'),
        AvgM=('margin_pct','mean'),Rank=('Rank','mean'),
        AvgNet=('Net price','mean'),AvgBB=('BB Price','mean'),
        FBA=('Number of FBA vendors','mean'),
    ).reset_index().round(1)
    bs['Rate']=(bs['Approved']/bs['Total']*100).round(1)
    bs=bs.sort_values('Approved',ascending=False)
    bs['Priority']=['🔥 IMMEDIATE' if i==0 else ('★ HIGH' if i<3 else 'PIPELINE') for i in range(len(bs))]

    r=3
    sub_hdr(ws,r,f"  BRAND RANKING BY APPROVED SKUs  ·  {len(bs)} brands",10); r+=1
    col_hdr(ws,r,[("Rank",6),("Brand",22),("Total",10),("Approved",11),("Rejected",10),
                   ("Rate",11),("Units/Mo",12),("Avg Margin%",13),("Avg Rank",11),
                   ("Priority",13)]); r+=1
    for ri,(_,row) in enumerate(bs.iterrows()):
        bg="FFF7ED" if ri==0 else ("FFFBEB" if ri<3 else (WH if ri%2==0 else LT))
        rf,rfont=rate_style(row['Rate'])
        vals=[ri+1,row['Brand'],int(row['Total']),int(row['Approved']),int(row['Rejected']),
              row['Rate'],int(row['Units']),row['AvgM'],round(row['Rank']),row['Priority']]
        for ci,v in enumerate(vals,1):
            c=ws.cell(row=r,column=ci,value=v)
            c.fill=rf if ci==6 else nf(bg); c.border=th(); c.alignment=cen()
            c.font=rfont if ci==6 else fd(10,b=(ri==0))
            if ci==8: c.font=fo(10)
        ws.row_dimensions[r].height=20; r+=1

    cht=BarChart(); cht.type="bar"; cht.title="Approved SKUs by Brand"
    cht.style=10; cht.width=18; cht.height=12
    de=Reference(ws,min_col=4,min_row=5,max_row=5+len(bs)-1)
    ct=Reference(ws,min_col=2,min_row=5,max_row=5+len(bs)-1)
    cht.add_data(de); cht.set_categories(ct)
    cht.series[0].graphicalProperties.solidFill=OR
    ws.add_chart(cht,f"A{r+2}")
    ws.freeze_panes="A5"
    ws.auto_filter.ref=f"A4:{get_column_letter(10)}{4+len(bs)}"

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 6 — REJECTION ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════
def sh_rejected(wb,d):
    ws=wb.create_sheet("⚠️ Rejection Analysis"); sg(ws)
    df=d['df']
    rej=df[~df['is_yes']]
    banner(ws,1,"REJECTION ANALYSIS — ROOT CAUSE & ACTION PLAN",10,32)
    ws.row_dimensions[2].height=8

    neg=int(rej['Remarks '].str.contains('Negative',na=False).sum()) if 'Remarks ' in rej.columns else 0
    low=int(rej['Remarks '].str.contains('Low',na=False).sum()) if 'Remarks ' in rej.columns else 0
    oth=len(rej)-neg-low

    r=3
    for i,(lbl,val,bg_h,fc) in enumerate([("TOTAL REJECTED",len(rej),LRD,RD),
                                            ("NEGATIVE MARGIN",neg,LRD,RD),
                                            ("LOW MARGIN",low,LAM,AM),
                                            ("OTHER",oth,LT,DK)]):
        c0=1+i*2
        ws.merge_cells(start_row=r,end_row=r,start_column=c0,end_column=c0+1)
        ws.merge_cells(start_row=r+1,end_row=r+2,start_column=c0,end_column=c0+1)
        cL=ws.cell(row=r,column=c0,value=lbl); cL.font=fd(8,c=fc); cL.fill=nf(bg_h); cL.alignment=lft()
        cV=ws.cell(row=r+1,column=c0,value=val); cV.font=Font(name="Calibri",bold=True,size=20,color=fc); cV.fill=nf(WH); cV.alignment=cen()
        ws.row_dimensions[r].height=14; ws.row_dimensions[r+1].height=22; ws.row_dimensions[r+2].height=8
    r+=3; ws.row_dimensions[r].height=10; r+=1

    sub_hdr(ws,r,"  REJECTED ASIN DETAIL",10); r+=1
    col_hdr(ws,r,[("ASIN",14),("Vendor",8),("Brand",16),("Analyst",11),
                   ("Net $",10),("BB $",10),("Breakeven",11),("Diff SP",10),
                   ("Margin$",10),("Score",9),("Reason",22)]); r+=1
    for _,row in rej.iterrows():
        gap=round(row.get('BB Price',0)-row.get('Net price',0),2)
        vals=[row.get('ASIN',''),row.get('Vendor Prefix',''),row.get('Brand',''),row.get('Analyst',''),
              round(row.get('Net price',0),2),round(row.get('BB Price',0),2),
              round(row.get('Breakeven',0),2),round(row.get('Difference from SP',0),2),
              gap,row.get('Total Product Score',''),str(row.get('Remarks ',''))]
        for ci,v in enumerate(vals,1):
            c=ws.cell(row=r,column=ci,value=v)
            c.fill=nf(LRD); c.border=th(); c.alignment=cen()
            c.font=Font(name="Calibri",size=9,color=RD if ci in (5,6,7,8) else DK,bold=(ci==3))
        ws.row_dimensions[r].height=18; r+=1

    r+=1
    sub_hdr(ws,r,"  MANAGEMENT ACTION PLAN",10); r+=1
    acts=[("1 — IMMEDIATE",f"{neg} ASINs: Negative Margin. BB Price < landed cost. Request 15-20% vendor price reduction before next PO. Do not order.",RD,LRD),
          ("2 — THIS WEEK",f"{low} ASINs: Low Margin. Explore volume pricing or bundles. Set minimum margin floor for future submissions.",AM,LAM),
          ("3 — NEXT CYCLE",f"All {len(rej)} rejections from PROH vendor. Review vendor relationship. Set margin thresholds before resubmission.",DK,LT)]
    for num,txt,fc,bg_h in acts:
        ws.merge_cells(f"A{r}:K{r}")
        c=ws.cell(row=r,column=1,value=f"  [{num}]  {txt}")
        c.font=Font(name="Calibri",size=10,color=fc,bold=(fc==RD))
        c.fill=nf(bg_h); c.alignment=lft(); ws.row_dimensions[r].height=24; r+=1

    ws.freeze_panes="A9"

# ═══════════════════════════════════════════════════════════════════════════
def generate_excel(data:dict)->bytes:
    wb=Workbook(); wb.remove(wb.active)
    sh_exec(wb,data)
    sh_vendor(wb,data)
    sh_asins(wb,data)
    sh_analysts(wb,data)
    sh_brands(wb,data)
    sh_rejected(wb,data)
    buf=io.BytesIO(); wb.save(buf)
    return buf.getvalue()
